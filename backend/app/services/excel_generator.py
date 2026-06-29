"""Generador de Excel de auditoría de riesgos contractuales v2.0"""
import json
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

# ── Paleta corporativa ────────────────────────────────────────────────────────
C_AZUL_OSC   = "1A365D"   # Encabezados principales
C_AZUL_MED   = "2B5797"   # Encabezados secundarios
C_AZUL_CLAR  = "D6E4F0"   # Filas alternas / fondo secciones
C_BLANCO     = "FFFFFF"
C_GRIS_CLAR  = "F1F5F9"
C_GRIS_BORD  = "CBD5E1"

# Criticidad
C_ALTA_BG    = "FECACA"   # rojo claro
C_ALTA_FG    = "991B1B"
C_MEDIA_BG   = "FED7AA"   # naranja claro
C_MEDIA_FG   = "92400E"
C_BAJA_BG    = "D1FAE5"   # verde claro
C_BAJA_FG    = "065F46"

# Tipos de riesgo
C_TIPO = {
    "LEGAL":          ("EDE9FE", "5B21B6"),
    "CONTRACTUAL":    ("FEE2E2", "991B1B"),
    "TECNICO":        ("DBEAFE", "1E40AF"),
    "ADMINISTRATIVO": ("FFEDD5", "9A3412"),
    "PROCEDIMENTAL":  ("D1FAE5", "065F46"),
}

# Precisión
C_PREC = {
    "MUY ALTA": ("D1FAE5", "065F46"),
    "ALTA":     ("D1FAE5", "065F46"),
    "MEDIA":    ("FEF3C7", "92400E"),
    "BAJA":     ("FECACA", "991B1B"),
    "MUY BAJA": ("FECACA", "991B1B"),
}

def _fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def _font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _border_thin():
    s = Side(style="thin", color=C_GRIS_BORD)
    return Border(left=s, right=s, top=s, bottom=s)
def _border_med():
    s = Side(style="medium", color=C_AZUL_MED)
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, row, col, texto, bg=C_AZUL_OSC, fg=C_BLANCO, size=10, bold=True, wrap=True):
    c = ws.cell(row=row, column=col, value=texto)
    c.font    = _font(bold=bold, color=fg, size=size)
    c.fill    = _fill(bg)
    c.alignment = _align("center", "center", wrap)
    c.border  = _border_thin()
    return c

def _cel(ws, row, col, value, bg=C_BLANCO, fg="000000", bold=False,
         wrap=True, h="left", size=9):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, color=fg, size=size)
    c.fill      = _fill(bg)
    c.alignment = _align(h, "center", wrap)
    c.border    = _border_thin()
    return c

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 1 — PORTADA
# ─────────────────────────────────────────────────────────────────────────────
def _hoja_portada(wb, data):
    ws = wb.active
    ws.title = "📋 Portada"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 3

    resultado = data.get("resultado") or {}
    metadata  = resultado.get("metadata_analisis") or {}
    riesgos   = resultado.get("riesgos_detectados") or []
    dist_tipo = metadata.get("distribucion_por_tipo") or {}
    dist_crit = metadata.get("distribucion_por_criticidad") or {}

    # Título
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value     = "INFORME DE AUDITORÍA CONTRACTUAL"
    c.font      = _font(bold=True, color=C_BLANCO, size=18)
    c.fill      = _fill(C_AZUL_OSC)
    c.alignment = _align("center", "center")

    ws.row_dimensions[3].height = 22
    ws.merge_cells("B3:C3")
    c = ws["B3"]
    c.value     = "Matriz de Riesgos — Análisis de Licitaciones y Contratos Públicos"
    c.font      = _font(italic=True, color=C_BLANCO, size=11)
    c.fill      = _fill(C_AZUL_MED)
    c.alignment = _align("center", "center")

    # Separador
    ws.row_dimensions[4].height = 8
    for col in ["B", "C"]:
        ws[f"{col}4"].fill = _fill(C_AZUL_CLAR)

    # Datos del análisis
    campos = [
        ("Licitación",         data.get("nombre_archivo_licitacion", "N/D")),
        ("Tipo de contrato",   data.get("tipo_contrato", "N/D")),
        ("Modalidad",          metadata.get("modalidad_ejecucion", "N/D")),
        ("Sector",             data.get("sector", "N/D")),
        ("Categoría",          data.get("categoria", "N/D")),
        ("Estado",             data.get("status", "N/D")),
        ("ID de análisis",     data.get("id_analisis", "N/D")),
        ("Fecha de emisión",   datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Normativas cargadas", ", ".join(data.get("normativas_cargadas", []))),
    ]

    row = 5
    for label, valor in campos:
        ws.row_dimensions[row].height = 20
        c_lab = ws.cell(row=row, column=2, value=label)
        c_lab.font      = _font(bold=True, color=C_AZUL_OSC, size=10)
        c_lab.fill      = _fill(C_AZUL_CLAR)
        c_lab.alignment = _align("right", "center")
        c_lab.border    = _border_thin()

        c_val = ws.cell(row=row, column=3, value=str(valor))
        c_val.font      = _font(size=10)
        c_val.fill      = _fill(C_BLANCO)
        c_val.alignment = _align("left", "center", wrap=True)
        c_val.border    = _border_thin()
        row += 1

    # Separador
    row += 1
    ws.row_dimensions[row].height = 8
    for col in [2, 3]:
        ws.cell(row=row, column=col).fill = _fill(C_AZUL_CLAR)

    # Resumen estadístico
    row += 1
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"B{row}:C{row}")
    c = ws.cell(row=row, column=2)
    c.value     = "RESUMEN ESTADÍSTICO"
    c.font      = _font(bold=True, color=C_BLANCO, size=11)
    c.fill      = _fill(C_AZUL_MED)
    c.alignment = _align("center", "center")
    row += 1

    total = len(riesgos)
    stats = [
        ("Total de riesgos detectados", total),
        ("Criticidad ALTA",  dist_crit.get("ALTA", sum(1 for r in riesgos if r.get("criticidad","").upper()=="ALTA"))),
        ("Criticidad MEDIA", dist_crit.get("MEDIA", sum(1 for r in riesgos if r.get("criticidad","").upper()=="MEDIA"))),
        ("Criticidad BAJA",  dist_crit.get("BAJA", sum(1 for r in riesgos if r.get("criticidad","").upper()=="BAJA"))),
        ("Requieren validación manual", sum(1 for r in riesgos if r.get("requiere_validacion_humana"))),
        ("Precisión MUY ALTA / ALTA",  sum(1 for r in riesgos if r.get("precision_nivel","") in ("MUY ALTA","ALTA"))),
    ]
    for label, val in stats:
        ws.row_dimensions[row].height = 20
        c_lab = ws.cell(row=row, column=2, value=label)
        c_lab.font = _font(bold=True, color=C_AZUL_OSC, size=10)
        c_lab.fill = _fill(C_GRIS_CLAR)
        c_lab.alignment = _align("right", "center")
        c_lab.border = _border_thin()

        c_val = ws.cell(row=row, column=3, value=val)
        c_val.font = _font(bold=True, size=11)
        c_val.fill = _fill(C_BLANCO)
        c_val.alignment = _align("center", "center")
        c_val.border = _border_thin()
        row += 1

    # Distribución por tipo
    if dist_tipo:
        row += 1
        ws.merge_cells(f"B{row}:C{row}")
        c = ws.cell(row=row, column=2)
        c.value = "DISTRIBUCIÓN POR TIPO"
        c.font = _font(bold=True, color=C_BLANCO, size=11)
        c.fill = _fill(C_AZUL_MED)
        c.alignment = _align("center", "center")
        ws.row_dimensions[row].height = 22
        row += 1

        for tipo, cant in dist_tipo.items():
            bg_t, fg_t = C_TIPO.get(tipo.upper(), (C_GRIS_CLAR, "000000"))
            ws.row_dimensions[row].height = 20
            c_t = ws.cell(row=row, column=2, value=tipo)
            c_t.font = _font(bold=True, color=fg_t, size=10)
            c_t.fill = _fill(bg_t)
            c_t.alignment = _align("center", "center")
            c_t.border = _border_thin()

            c_v = ws.cell(row=row, column=3, value=int(cant) if str(cant).isdigit() else cant)
            c_v.font = _font(bold=True, size=10)
            c_v.fill = _fill(C_BLANCO)
            c_v.alignment = _align("center", "center")
            c_v.border = _border_thin()
            row += 1

    # Análisis macro viabilidad
    macro = resultado.get("analisis_macro_viabilidad") or resultado.get("resumen_ejecutivo", "")
    if macro:
        row += 1
        ws.merge_cells(f"B{row}:C{row}")
        c = ws.cell(row=row, column=2)
        c.value = "ANÁLISIS DE VIABILIDAD"
        c.font = _font(bold=True, color=C_BLANCO, size=11)
        c.fill = _fill(C_AZUL_MED)
        c.alignment = _align("center", "center")
        ws.row_dimensions[row].height = 22
        row += 1

        ws.row_dimensions[row].height = 80
        ws.merge_cells(f"B{row}:C{row}")
        c = ws.cell(row=row, column=2, value=macro)
        c.font = _font(size=10, italic=True, color=C_AZUL_OSC)
        c.fill = _fill(C_AZUL_CLAR)
        c.alignment = _align("left", "top", wrap=True)
        c.border = _border_med()

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 2 — MATRIZ COMPLETA (tabla principal de auditoría)
# ─────────────────────────────────────────────────────────────────────────────
def _hoja_matriz(wb, data):
    ws = wb.create_sheet("📊 Matriz de Riesgos")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"  # Congelar primeras 2 filas

    resultado = data.get("resultado") or {}
    riesgos   = resultado.get("riesgos_detectados") or []

    # Columnas: (encabezado, ancho)
    cols = [
        ("N°",                           6),
        ("Tipo",                         16),
        ("Origen Matriz",                16),
        ("ID Matriz",                    10),
        ("Riesgo Identificado",          45),
        ("Criticidad",                   12),
        ("Sección / Cláusula",           22),
        ("Evidencia Licitación",         55),
        ("Sustento Legal / Normativo",   45),
        ("Impacto Comercial/Financiero", 50),
        ("Propuesta de Observación Formal", 60),
        ("Precisión",                    12),
        ("Score",                        8),
        ("Validación Manual",            16),
        ("Evidencia Localizada",         16),
        ("Sustento Localizado",          16),
    ]

    # Fila 1: título de la hoja
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    c = ws["A1"]
    c.value     = "MATRIZ DE RIESGOS CONTRACTUALES — AUDITORÍA DE LICITACIÓN"
    c.font      = _font(bold=True, color=C_BLANCO, size=13)
    c.fill      = _fill(C_AZUL_OSC)
    c.alignment = _align("center", "center")

    # Fila 2: encabezados de columnas
    ws.row_dimensions[2].height = 36
    for ci, (nombre, ancho) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = ancho
        _hdr(ws, 2, ci, nombre, bg=C_AZUL_MED, size=9)

    # Datos
    for ri, r in enumerate(riesgos, start=1):
        row = ri + 2
        ws.row_dimensions[row].height = 90

        criticidad = str(r.get("criticidad", "")).upper()
        tipo       = str(r.get("tipo_riesgo", "")).upper()
        precision  = str(r.get("precision_nivel", "")).upper()

        bg_crit = {"ALTA": C_ALTA_BG, "MEDIA": C_MEDIA_BG, "BAJA": C_BAJA_BG}.get(criticidad, C_BLANCO)
        fg_crit = {"ALTA": C_ALTA_FG, "MEDIA": C_MEDIA_FG, "BAJA": C_BAJA_FG}.get(criticidad, "000000")
        bg_tipo, fg_tipo = C_TIPO.get(tipo, (C_GRIS_CLAR, "000000"))
        bg_prec, fg_prec = C_PREC.get(precision, (C_BLANCO, "000000"))

        # Fila base alterna
        bg_row = C_GRIS_CLAR if ri % 2 == 0 else C_BLANCO

        traz    = r.get("trazabilidad") or {}
        traz_l  = traz.get("evidencia_licitacion") or {}
        traz_n  = traz.get("sustento_legal_normativo") or {}
        ev_ok   = "✓ Sí" if traz_l.get("encontrado") else "✗ No"
        sus_ok  = "✓ Sí" if traz_n.get("encontrado") else "✗ No"
        val_hum = "⚠ SÍ" if r.get("requiere_validacion_humana") else "✓ No"

        valores = [
            (r.get("numero_riesgo", ri),       bg_row,   "000000", True,  "center"),
            (tipo,                              bg_tipo,  fg_tipo,  True,  "center"),
            (r.get("origen_matriz", ""),        bg_row,   "000000", False, "center"),
            (r.get("id_riesgo_matriz", ""),     bg_row,   "000000", False, "center"),
            (r.get("riesgo_identificado", ""),  bg_row,   C_AZUL_OSC, True, "left"),
            (criticidad,                        bg_crit,  fg_crit,  True,  "center"),
            (r.get("seccion_bases", ""),        bg_row,   "000000", False, "left"),
            (r.get("evidencia_licitacion", ""), C_GRIS_CLAR, "334155", False, "left"),
            (r.get("sustento_legal_normativo", ""), C_GRIS_CLAR, "334155", False, "left"),
            (r.get("impacto_comercial_financiero", ""), bg_row, "000000", False, "left"),
            (r.get("propuesta_observacion_formal", ""), "EFF6FF", "1E3A5F", False, "left"),
            (precision,                         bg_prec,  fg_prec,  True,  "center"),
            (r.get("precision_score", 0),       bg_prec,  fg_prec,  True,  "center"),
            (val_hum,                           "FEF3C7" if r.get("requiere_validacion_humana") else "D1FAE5",
                                                "92400E" if r.get("requiere_validacion_humana") else "065F46",
                                                True, "center"),
            (ev_ok,  "D1FAE5" if traz_l.get("encontrado") else "FECACA",
                     "065F46" if traz_l.get("encontrado") else "991B1B", True, "center"),
            (sus_ok, "D1FAE5" if traz_n.get("encontrado") else "FECACA",
                     "065F46" if traz_n.get("encontrado") else "991B1B", True, "center"),
        ]

        for ci, (val, bg, fg, bold, h) in enumerate(valores, start=1):
            _cel(ws, row, ci, val, bg=bg, fg=fg, bold=bold, wrap=True, h=h)

    # Tabla nativa de Excel (filtros automáticos)
    if riesgos:
        last_row = len(riesgos) + 2
        last_col = get_column_letter(len(cols))
        tabla = Table(
            displayName="MatrizRiesgos",
            ref=f"A2:{last_col}{last_row}",
        )
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tabla)

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 3 — DETALLE DE TRAZABILIDAD
# ─────────────────────────────────────────────────────────────────────────────
def _hoja_trazabilidad(wb, data):
    ws = wb.create_sheet("🔍 Trazabilidad")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    resultado = data.get("resultado") or {}
    riesgos   = resultado.get("riesgos_detectados") or []

    cols = [
        ("N°",                        6),
        ("Riesgo",                    40),
        ("Criticidad",                12),
        ("Ev. Licitación — Fragmento literal",  55),
        ("Ev. Licitación — Posición (car.)",    16),
        ("Ev. Licitación — ¿Localizada?",       16),
        ("Sustento Legal — Fragmento literal",  55),
        ("Sustento Legal — Posición (car.)",    16),
        ("Sustento Legal — ¿Localizado?",       16),
        ("Referencias No Verificadas",          30),
        ("Alerta del Sistema",                  40),
        ("Detalle de Precisión",                55),
    ]

    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    c = ws["A1"]
    c.value     = "DETALLE DE TRAZABILIDAD Y VERIFICACIÓN DE FUENTES"
    c.font      = _font(bold=True, color=C_BLANCO, size=13)
    c.fill      = _fill(C_AZUL_OSC)
    c.alignment = _align("center", "center")

    ws.row_dimensions[2].height = 36
    for ci, (nombre, ancho) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = ancho
        _hdr(ws, 2, ci, nombre, bg=C_AZUL_MED, size=9)

    for ri, r in enumerate(riesgos, start=1):
        row = ri + 2
        ws.row_dimensions[row].height = 80

        criticidad = str(r.get("criticidad", "")).upper()
        bg_crit = {"ALTA": C_ALTA_BG, "MEDIA": C_MEDIA_BG, "BAJA": C_BAJA_BG}.get(criticidad, C_BLANCO)
        fg_crit = {"ALTA": C_ALTA_FG, "MEDIA": C_MEDIA_FG, "BAJA": C_BAJA_FG}.get(criticidad, "000000")
        bg_row = C_GRIS_CLAR if ri % 2 == 0 else C_BLANCO

        traz   = r.get("trazabilidad") or {}
        traz_l = traz.get("evidencia_licitacion") or {}
        traz_n = traz.get("sustento_legal_normativo") or {}

        frag_l = traz_l.get("fragmento_literal_fuente") or r.get("evidencia_licitacion", "")
        pos_l  = traz_l.get("indice_aproximado_caracter", "N/D")
        ok_l   = "✓ Sí" if traz_l.get("encontrado") else "✗ No"

        frag_n = traz_n.get("fragmento_literal_fuente") or r.get("sustento_legal_normativo", "")
        pos_n  = traz_n.get("indice_aproximado_caracter", "N/D")
        ok_n   = "✓ Sí" if traz_n.get("encontrado") else "✗ No"

        refs_nv = ", ".join(traz_n.get("referencias_no_localizadas", []))
        alerta  = r.get("alerta_sistema") or ""
        detalle = "\n".join(r.get("precision_detalle") or [])

        filas_vals = [
            (r.get("numero_riesgo", ri),   bg_row,   "000000", True,  "center"),
            (r.get("riesgo_identificado", ""), bg_row, C_AZUL_OSC, True, "left"),
            (criticidad,                   bg_crit,  fg_crit,  True,  "center"),
            (frag_l,                       C_GRIS_CLAR, "334155", False, "left"),
            (pos_l,                        bg_row,   "000000", False, "center"),
            (ok_l,  "D1FAE5" if traz_l.get("encontrado") else "FECACA",
                    "065F46" if traz_l.get("encontrado") else "991B1B", True, "center"),
            (frag_n,                       C_GRIS_CLAR, "334155", False, "left"),
            (pos_n,                        bg_row,   "000000", False, "center"),
            (ok_n,  "D1FAE5" if traz_n.get("encontrado") else "FECACA",
                    "065F46" if traz_n.get("encontrado") else "991B1B", True, "center"),
            (refs_nv, "FEF3C7" if refs_nv else bg_row, "92400E" if refs_nv else "000000", False, "left"),
            (alerta,  "FEF3C7" if alerta else bg_row, "92400E" if alerta else "000000", False, "left"),
            (detalle, bg_row, "000000", False, "left"),
        ]
        for ci, (val, bg, fg, bold, h) in enumerate(filas_vals, start=1):
            _cel(ws, row, ci, val, bg=bg, fg=fg, bold=bold, wrap=True, h=h)

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 4 — OBSERVACIONES FORMALES (pliego listo para enviar)
# ─────────────────────────────────────────────────────────────────────────────
def _hoja_observaciones(wb, data):
    ws = wb.create_sheet("📝 Observaciones Formales")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    resultado = data.get("resultado") or {}
    riesgos   = resultado.get("riesgos_detectados") or []

    cols = [
        ("N°",                  6),
        ("Sección / Cláusula", 25),
        ("Tipo",               16),
        ("Criticidad",         12),
        ("Riesgo Identificado",40),
        ("Propuesta de Observación Formal al Pliego", 80),
        ("Estado",             16),
    ]

    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    c = ws["A1"]
    c.value     = "PLIEGO DE CONSULTAS Y OBSERVACIONES — LISTAS PARA ENVÍO A MESA DE PARTES"
    c.font      = _font(bold=True, color=C_BLANCO, size=13)
    c.fill      = _fill(C_AZUL_OSC)
    c.alignment = _align("center", "center")

    ws.row_dimensions[2].height = 36
    for ci, (nombre, ancho) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = ancho
        _hdr(ws, 2, ci, nombre, bg=C_AZUL_MED, size=9)

    # Solo riesgos con propuesta de observación
    riesgos_con_obs = [r for r in riesgos if r.get("propuesta_observacion_formal")]
    for ri, r in enumerate(riesgos_con_obs, start=1):
        row = ri + 2
        ws.row_dimensions[row].height = 100

        criticidad = str(r.get("criticidad", "")).upper()
        tipo       = str(r.get("tipo_riesgo", "")).upper()
        bg_crit = {"ALTA": C_ALTA_BG, "MEDIA": C_MEDIA_BG, "BAJA": C_BAJA_BG}.get(criticidad, C_BLANCO)
        fg_crit = {"ALTA": C_ALTA_FG, "MEDIA": C_MEDIA_FG, "BAJA": C_BAJA_FG}.get(criticidad, "000000")
        bg_tipo, fg_tipo = C_TIPO.get(tipo, (C_GRIS_CLAR, "000000"))
        bg_row = C_GRIS_CLAR if ri % 2 == 0 else C_BLANCO

        vals = [
            (r.get("numero_riesgo", ri),              bg_row,  "000000", True,  "center"),
            (r.get("seccion_bases", ""),               bg_row,  C_AZUL_OSC, False, "left"),
            (tipo,                                     bg_tipo, fg_tipo,  True,  "center"),
            (criticidad,                               bg_crit, fg_crit,  True,  "center"),
            (r.get("riesgo_identificado", ""),         bg_row,  C_AZUL_OSC, True, "left"),
            (r.get("propuesta_observacion_formal", ""),"EFF6FF","1E3A5F", False, "left"),
            ("PENDIENTE",                              "FEF3C7","92400E", True,  "center"),
        ]
        for ci, (val, bg, fg, bold, h) in enumerate(vals, start=1):
            _cel(ws, row, ci, val, bg=bg, fg=fg, bold=bold, wrap=True, h=h)

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 5 — DASHBOARD RESUMEN por tipo y criticidad
# ─────────────────────────────────────────────────────────────────────────────
def _hoja_dashboard(wb, data):
    ws = wb.create_sheet("📈 Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 3

    resultado = data.get("resultado") or {}
    riesgos   = resultado.get("riesgos_detectados") or []

    tipos      = ["LEGAL","CONTRACTUAL","TECNICO","ADMINISTRATIVO","PROCEDIMENTAL"]
    crits      = ["ALTA","MEDIA","BAJA"]
    precisiones = ["MUY ALTA","ALTA","MEDIA","BAJA","MUY BAJA"]

    # Calcular tabla cruzada: tipo × criticidad
    tabla = {t: {c: 0 for c in crits} for t in tipos}
    for r in riesgos:
        t = str(r.get("tipo_riesgo","")).upper()
        c = str(r.get("criticidad","")).upper()
        if t in tabla and c in crits:
            tabla[t][c] += 1

    prec_counts = {p: 0 for p in precisiones}
    for r in riesgos:
        p = str(r.get("precision_nivel","")).upper()
        if p in prec_counts:
            prec_counts[p] += 1

    row = 2
    ws.row_dimensions[row].height = 28
    ws.merge_cells(f"B{row}:F{row}")
    c = ws.cell(row=row, column=2)
    c.value     = "DASHBOARD — DISTRIBUCIÓN DE RIESGOS"
    c.font      = _font(bold=True, color=C_BLANCO, size=14)
    c.fill      = _fill(C_AZUL_OSC)
    c.alignment = _align("center", "center")
    row += 2

    # Tabla cruzada tipo × criticidad
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"B{row}:F{row}")
    c = ws.cell(row=row, column=2)
    c.value = "Riesgos por Tipo y Criticidad"
    c.font  = _font(bold=True, color=C_BLANCO, size=11)
    c.fill  = _fill(C_AZUL_MED)
    c.alignment = _align("center", "center")
    row += 1

    # Encabezados
    _hdr(ws, row, 2, "Tipo de Riesgo", bg=C_AZUL_MED, size=10)
    for ci, crit in enumerate(crits, start=3):
        bg = {"ALTA": C_ALTA_BG, "MEDIA": C_MEDIA_BG, "BAJA": C_BAJA_BG}[crit]
        fg = {"ALTA": C_ALTA_FG, "MEDIA": C_MEDIA_FG, "BAJA": C_BAJA_FG}[crit]
        _hdr(ws, row, ci, crit, bg=bg, fg=fg, size=10)
    _hdr(ws, row, 6, "TOTAL", bg=C_AZUL_OSC, size=10)
    row += 1

    for tipo in tipos:
        ws.row_dimensions[row].height = 20
        bg_t, fg_t = C_TIPO.get(tipo, (C_GRIS_CLAR, "000000"))
        _cel(ws, row, 2, tipo, bg=bg_t, fg=fg_t, bold=True, h="center")
        total_tipo = 0
        for ci, crit in enumerate(crits, start=3):
            val = tabla[tipo][crit]
            total_tipo += val
            bg = {"ALTA": C_ALTA_BG, "MEDIA": C_MEDIA_BG, "BAJA": C_BAJA_BG}[crit]
            _cel(ws, row, ci, val, bg=bg if val > 0 else C_BLANCO,
                 fg={"ALTA": C_ALTA_FG, "MEDIA": C_MEDIA_FG, "BAJA": C_BAJA_FG}[crit] if val > 0 else "AAAAAA",
                 bold=val > 0, h="center")
        _cel(ws, row, 6, total_tipo, bg=C_AZUL_CLAR, fg=C_AZUL_OSC, bold=True, h="center")
        row += 1

    # Fila de totales
    ws.row_dimensions[row].height = 22
    _hdr(ws, row, 2, "TOTAL", bg=C_AZUL_OSC, size=10)
    for ci, crit in enumerate(crits, start=3):
        total_col = sum(tabla[t][crit] for t in tipos)
        _hdr(ws, row, ci, total_col, bg=C_AZUL_OSC, size=10)
    _hdr(ws, row, 6, len(riesgos), bg=C_AZUL_OSC, size=11)
    row += 3

    # Tabla de precisión
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"B{row}:F{row}")
    c = ws.cell(row=row, column=2)
    c.value = "Distribución por Nivel de Precisión de Detección"
    c.font  = _font(bold=True, color=C_BLANCO, size=11)
    c.fill  = _fill(C_AZUL_MED)
    c.alignment = _align("center", "center")
    row += 1

    _hdr(ws, row, 2, "Nivel de Precisión", bg=C_AZUL_MED, size=10)
    _hdr(ws, row, 3, "Cantidad", bg=C_AZUL_MED, size=10)
    _hdr(ws, row, 4, "% del Total", bg=C_AZUL_MED, size=10)
    row += 1

    total_r = len(riesgos) or 1
    for prec in precisiones:
        ws.row_dimensions[row].height = 20
        cnt = prec_counts[prec]
        bg_p, fg_p = C_PREC.get(prec, (C_BLANCO, "000000"))
        _cel(ws, row, 2, prec, bg=bg_p, fg=fg_p, bold=True, h="center")
        _cel(ws, row, 3, cnt, bg=bg_p if cnt > 0 else C_BLANCO,
             fg=fg_p if cnt > 0 else "AAAAAA", bold=cnt > 0, h="center")
        pct_val = round(cnt / total_r * 100, 1)
        c_pct = ws.cell(row=row, column=4, value=f"{pct_val}%")
        c_pct.font = _font(size=10, color=fg_p if cnt > 0 else "AAAAAA")
        c_pct.fill = _fill(bg_p if cnt > 0 else C_BLANCO)
        c_pct.alignment = _align("center", "center")
        c_pct.border = _border_thin()
        row += 1

# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────
def generar_excel_informe(data_analisis: dict) -> bytes:
    wb = Workbook()

    _hoja_portada(wb, data_analisis)
    _hoja_matriz(wb, data_analisis)
    _hoja_trazabilidad(wb, data_analisis)
    _hoja_observaciones(wb, data_analisis)
    _hoja_dashboard(wb, data_analisis)

    # Activar la portada al abrir
    wb.active = wb["📋 Portada"]

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Test con el JSON de muestra ───────────────────────────────────────────────